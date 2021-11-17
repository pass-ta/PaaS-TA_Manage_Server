# from typing import AwaitableGenerator
from django.shortcuts import render
from django.shortcuts import redirect, render
from django.core.files.storage import FileSystemStorage
from home import models
from main.models import User
import string
import random
from django.core.paginator import Paginator, EmptyPage

from home.models import Enrol, Notice, Room, Analytics, SolvedQuiz
import string
import random

from home.models import Quiz, Room
import os
import base64
from luxand import luxand
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from HiClass import settings
from .fusioncharts import FusionCharts

from django.db.models.fields import NullBooleanField
from django.db.models.query import QuerySet
from django.http import response
from django.urls.conf import path
from django.views.generic import ListView
from collections import OrderedDict
# from .fusioncharts import FusionCharts
import simplejson
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http.response import HttpResponse, JsonResponse
from django.core import serializers
from django.contrib import messages

import os
import base64


def home(request):
    res_data = {}
    user_session = request.session.get('user')              # 로그인 체크
    fs = FileSystemStorage()
    if user_session:
        user = User.objects.get(pk=user_session)
        res_data['username'] = user.username                # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        print(user.image)
        res_data['userimg'] = fs.url(user.image)

        print(res_data['userimg'], "!!!!!!!!!!!!!!!!!!!!")

        if res_data['userimg'] == "/media/face-recognition.png":               # 이미지 체크
            res_data['img_check'] = 0                      # 이미지 널
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            if user.role == 'student':
                return render(request, 'home-s.html', res_data)
            else:
                return render(request, 'home.html', res_data)
        elif request.method == 'POST':
            userimage = request.FILES['user-img-change']
            res_data['userimg'] = fs.url(userimage)
            user.image = userimage
            user.save()
            res_data['img_check'] = 1
            if user.role == 'student':
                return render(request, 'home-s.html', res_data)
            else:
                return render(request, 'home.html', res_data)
    else:
        return redirect('/login')


def makeclass(request):
    res_data = {}
    user_session = request.session.get('user')
    fs = FileSystemStorage()
    if user_session:
        user = User.objects.get(pk=user_session)            # 로그인 체크
        res_data['username'] = user.username                # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['string'] = ''.join(random.choice(
            string.ascii_uppercase + string.digits)for _ in range(7))  # 랜덤 문자열 생성
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지가 있는지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'makeclass.html', res_data)
        elif request.method == 'POST':
            room_id = request.POST.get('room-id', None)
            room_password = request.POST.get('room-password', None)
            room_name = request.POST.get('room-name', None)
            maker = user.email
            member_list = []

            try:
                file = request.FILES['file']
            except:
                file = "NULL"

            # 학생명단 file
            # 명단에서학번만 추출
            if not(file == "NULL"):
                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                member = load_workbook("media/" + file.name)
                for cell in member['Sheet1']['A']:
                    member_list.append(cell.value)

                os.remove(os.path.join(settings.MEDIA_ROOT, file.name))

            if not(room_password):
                res_data['password_error'] = '비밀번호를 생성해 주세요.'
            elif (not(room_name)):
                res_data['name_error'] = 'Class의 이름을 적어 주세요.'
            elif (file == "NULL"):
                res_data['file_error'] = '출석부를 첨부 해주세요.'
            else:
                room = Room(room_id=room_id, room_password=room_password, room_name=room_name,
                            file=file, maker=maker, member_list=member_list)  # db에 room 정보 저장
                room.save()
                # 방을 성공적으로 만들면 room_name으로 room_session을 저장
                request.session['room_id'] = room_id
                return redirect('/home/makeclass/success')
            # room 정보 비정상 일시
            return render(request, 'makeclass.html', res_data)
    else:
        return redirect('/login')


def make_success(request):
    res_data = {}
    fs = FileSystemStorage()
    room_session = request.session.get(
        'room_id')   # 아까 POST 할때 session에 저장한 값 불러옴
    user_session = request.session.get('user')
    if room_session and user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date

        # 가장 최근의 room_name과 session에 저장한 것을 비교함
        room = Room.objects.get(room_id=room_session)
        res_data['room_id'] = room.room_id
        res_data['room_name'] = room.room_name
        res_data['room_password'] = room.room_password
        res_data['maker'] = room.maker

        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1
        if request.method == 'GET':
            return render(request, 'make_success.html', res_data)
        elif request.method == 'POST':
            if user.role == "student":
                return redirect('/home/enterclass/student1')
            elif user.role == "teacher":
                return redirect('/home/enterclass/teacher')
    else:
        return redirect('/login')


def enterclass(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enterclass.html', res_data)
        elif request.method == 'POST':
            room_id = request.POST.get('room_id')
            room_password = request.POST.get('room_password')

            if not(room_id):
                res_data['name_error'] = 'Class ID를 입력하세요.'
            elif not(room_password):
                res_data['password_error'] = 'Class 비밀번호를 입력하세요.'
            elif not(room_id and room_password):
                res_data['all_error'] = '모든 값을 입력하세요.'
            else:
                try:
                    # 필드명 = 값 이면 Room 객체 생성
                    room = Room.objects.get(room_id=room_id)
                except Room.DoesNotExist:
                    # room이 없는 예외 처리
                    res_data['error'] = '존재하지 않는 Class 입니다.'
                    return render(request, 'enterclass.html', res_data)

                # 방 입장하는 순간 room session의 기준은 입장한 방 이름
                request.session['room_id'] = room_id
                db_password = room.room_password
                if db_password == room_password:     # room 정상 입장
                    if user.role == 'student':
                        return redirect('/home/enterclass/student1')
                    elif user.role == 'teacher':
                        return redirect('/home/class/makequiz')
                else:
                    res_data['error'] = '비밀번호가 틀렸습니다.'
                    return render(request, 'enterclass.html', res_data)
            # room 정보 비정상 일시
            return render(request, 'enterclass.html', res_data)
    else:
        return redirect('/login')


@method_decorator(csrf_exempt, name='dispatch')
def teacher(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    res_data['session'] = user_session
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enter_teacher.html', res_data)
        elif request.method == 'POST':
            room_session = request.session.get('room_id')
            room = Room.objects.get(room_id=room_session)
            roomid = room.room_id
            roomname = room.room_name
            useremail = user.email
            roomowner = room.maker
            nickname = user.username
            url = 'https://pedantic-einstein-75bdbe.netlify.app/'+roomid + \
                '/'+useremail+'/'+roomowner+'/'+nickname+'/'+"teacher"
            return redirect(url)
    else:
        return redirect('/login')


def student1(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        # room을 만든 사람 username 가져오기
        room_session = request.session.get('room_id')
        room = Room.objects.get(room_id=room_session)
        room_maker = User.objects.get(email=room.maker)
        res_data['room_maker'] = room_maker.username

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enter_student1.html', res_data)
        elif request.method == 'POST':
            # if user.check == True:
            return redirect('/home/enterclass/student2')
            # else:
            #res_data['check'] = "차단이 완료되지 않았습니다."
            # return render(request, 'enter_student1.html', res_data)
    else:
        return redirect('/login')


@csrf_exempt
def student2(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        print(user.username)
        # room을 만든 사람 username 가져오기
        room_session = request.session.get('room_id')
        room = Room.objects.get(room_id=room_session)
        room_maker = User.objects.get(email=room.maker)
        res_data['room_maker'] = room_maker.username

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':

            return render(request, 'enter_student2.html', res_data)

        elif request.method == 'POST':
            info = {}
            room_id = request.session.get('room_id')
            member_number = request.POST.get('member_number')
            member_name = request.POST.get('member_name')

            # 해당방의 DB속 명단Excel파일 조회
            try:
                room = Room.objects.get(room_id=room_id)
            except Room.DoesNotExist:
                return messages.warning(request, '존재 하는 Class가 없습니다.')
            member_file = room.file  # 명단

            # DB의 member_list로 회원번호 확인 및 index 추출
            member_list = room.member_list  # 회원번호만 적힌 리스트
            member_list = member_list[1:-1].split(', ')

            # CHECK NUMBER
            # Correct NUMBER
            if (member_number in member_list):
                # index=0은'회원번호(수험번호/학번)'이므로 index로 추출된 수 +1로 쓰면됨!
                member_index = member_list.index(member_number) + 1
                member = load_workbook("media/" + str(member_file))
                sheet = member['Sheet1']
                member_file_name = sheet['B'+str(member_index)].value

                # CHECK NAME
                # Wrong NAME
                if member_file_name != member_name:
                    info['result'] = "NO_NAME"
                # Correct NAME
                else:
                    # WEB : 캡쳐이미지 받기
                    member_image_data = request.POST.__getitem__('photo')
                    member_image_data = member_image_data[22:]
                    member_image_path = str(
                        room_id)+'_'+str(member_number)+'_capture.png'
                    member_image = open(os.path.join(
                        FileSystemStorage().location)+str("/capture/")+member_image_path, "wb")
                    member_image.write(base64.b64decode(member_image_data))
                    member_image.close()

                    # exel 명단 속 이미지
                    image_loader = SheetImageLoader(sheet)
                    image = image_loader.get('C'+str(member_index))

                    member_file_image_path = (
                        room_id+"_"+str(member_number)+".jpg")
                    image.save("media/capture/"+member_file_image_path)
                    fs = FileSystemStorage()

                    # Face Recognition
                    a = (fs.location + str("/capture/") + member_file_image_path)
                    b = (fs.location + str("/capture/") + member_image_path)

                    # luxand API
                    luxand_client = luxand("eed3a1c052394c12ac437d78651522f6")
                    member_file_image = luxand_client.add_person(
                        str(member_file_name), photos=[a])
                    result = luxand_client.verify(member_file_image, photo=b)

                    # Recognition RESULT
                    if result['status'] == 'success':
                        info['result'] = "OK"
                        print("Recognition_SUCCESS")
                    else:
                        info['result'] = "NO_IMAGE_MATCH"
                        print("Recognition_FAIL")
            # Wrong NUMBER
            else:
                # 명단 속 존재하지 않는 회원번호 (입장불가!)
                info['result'] = "NO_MEMBER"
                print('no_member')
                # 해당페이지에서 팝업으로 입장불가표시주기
            return JsonResponse(info)
    else:
        return redirect('/login')


def student3(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role

        # room을 만든 사람 username 가져오기
        room_session = request.session.get('room_id')
        room = Room.objects.get(room_id=room_session)
        room_maker = User.objects.get(email=room.maker)
        res_data['room_maker'] = room_maker.username

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        # room = Room(room_id=room_id,room_password=room_password,room_name=room_name,
        #                     file=file,maker=maker, member_list = member_list)  # db에 room 정보 저장
        #         room.save()

        if request.method == 'GET':
            return render(request, 'enter_teacher.html', res_data)
        elif request.method == 'POST':
            room_session = request.session.get('room_id')
            room = Room.objects.get(room_id=room_session)

            try:
                enrol = Enrol.objects.get(email=user.email, room_id=room.room_id)  # 수강 목록을 살펴 보고
            
            except Enrol.DoesNotExist:                                             # 없으면 enrol에 추가(중복 방지)
                enrol = Enrol(email=user.email, room_id=room.room_id, room_password=room.room_password,room_name=room.room_name)
                enrol.save()

            roomid = room.room_id
            roomname = room.room_name
            useremail = user.email
            roomowner = room.maker
            nickname = user.username

            url = 'https://pedantic-einstein-75bdbe.netlify.app/'+roomid + \
                '/'+useremail+'/'+roomowner+'/'+nickname+'/'+"student"
            # 룸 네임, user email, room owner , nickname, string
            return redirect(url)
    else:
        return redirect('/login')


class ClassList_t(ListView):
    model = Room
    template_name = 'myClass_t.html'

    def get_queryset(self):    # roomlist를 보여줄 queryset 특정
        # session에 저장되어 있는 email과 room의 maker가 같은 것만 queryset에 넣음
        QuerySet = Room.objects.filter(
            maker=self.request.session.get('user_email')).order_by('-make_date')
        return QuerySet


def classDetail_t(request, pk):
    room = Room.objects.get(pk=pk)
    request.session['room_id'] = room.room_id
    res_data = {}
    res_data['room_pk'] = room.pk
    res_data['room_name'] = room.room_name

    page = request.GET.get("page",1)
    notice_list = models.Notice.objects.filter(room_id = room.room_id).order_by('-make_date')
    paginator = Paginator(notice_list,100,orphans=5)
    try:
        notice = paginator.page(int(page))
    except EmptyPage:
        pass
    res_data["page"] = notice

    if request.method == 'GET':
        return render(request, 'classDetail-t.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail-t.html', res_data)

def classDetail_Detail_t(request,pk,pkk):
    room = Room.objects.get(pk = pk)
    request.session['room_id'] = room.room_id
    res_data = {}
    res_data['room_pk'] = room.pk
    res_data['room_name'] = room.room_name

    notice = Notice.objects.get(pk =pkk, room_id = room.room_id)
    res_data['title'] = notice.title
    res_data['description'] = notice.description
    res_data['writername'] = notice.writername
    res_data['date'] = notice.make_date

    if request.method == 'GET':
        return render(request,'classDetail-Detail-t.html',res_data)
    elif request.method == 'POST':
        return  render(request,'classDetail-Detail-t.html',res_data)

def makeNotice(request,pk):
    # room_session = request.session.get('room_id')
    room = Room.objects.get(pk=pk)
    user_session = request.session.get('user')
    user = User.objects.get(pk=user_session)

    res_data = {}
    res_data['room_pk'] = room.pk
    res_data['room_name'] = room.room_name
    
    if request.method == 'GET':
        return render(request,'makenotice.html',res_data)
    elif request.method == 'POST':
        title = request.POST.get('title', None)
        description = request.POST.get('description', None)

        notice = Notice(room_id=room.room_id, writer=user.email,
                        writername=user.username, title=title, description=description)
        notice.save()
        return redirect('/home/myclass/teacher/'+str(pk))


def classDetail2_t(request):
    room_session = request.session.get('room_id')
    room = Room.objects.get(room_id=room_session)
    res_data={}

    page = request.GET.get("page",1)
    quiz_list = models.Quiz.objects.filter(room_id = room.room_id).order_by('-make_date')
    paginator = Paginator(quiz_list,100,orphans=5)
    try:
        quiz = paginator.page(int(page))
    except EmptyPage:
        pass
    res_data["page"] = quiz
    for i in quiz_list:
        print(i.question,"!!!!!!!!!!!!!!!!!!!!!")

    res_data['room_pk'] = room.pk
    res_data['room_name'] = room.room_name
    if request.method == 'GET':
        return render(request, 'classDetail2-t.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail2-t.html', res_data)

def classDetail2_Detail2_t(request,pk):
    room_session = request.session.get('room_id')
    room = Room.objects.get(room_id = room_session)
    res_data = {}
    res_data['room_pk'] = room.pk
    res_data['room_name'] = room.room_name

    quiz = Quiz.objects.get(pk =pk, room_id = room.room_id)
    res_data['makername'] = quiz.makername
    res_data['question'] = quiz.question
    res_data['item1'] = quiz.item1
    res_data['item2'] = quiz.item2
    res_data['item3'] = quiz.item3
    res_data['item4'] = quiz.item4
    res_data['answer'] = quiz.answer
    res_data['date'] = quiz.make_date

    if request.method == 'GET':
        return render(request,'classDetail2-Detail2-t.html',res_data)
    elif request.method == 'POST':
        return  render(request,'classDetail2-Detail2-t.html',res_data)

def classDetail3_t(request):
    room_session = request.session.get('room_id')
    user_session = request.session.get('user')
    user = User.objects.get(pk=user_session)
    room = Room.objects.get(room_id=room_session)

    res_data = {}
    res_data['room_pk'] = room.pk
    res_data['room_name'] = room.room_name

    page = request.GET.get("page", 1)
    analytics_list = models.Analytics.objects.filter(
        room_id=room.room_id).order_by('username')
    paginator = Paginator(analytics_list, 10, orphans=5)
    try:
        analytics = paginator.page(int(page))
    except EmptyPage:
        pass
    res_data["page"] = analytics
    if request.method == 'GET':
        return render(request, 'classDetail3-t.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail3-t.html', res_data)


class ClassList_s(ListView):
    model = Enrol
    template_name = 'myClass_s.html'

    def get_queryset(self):    # roomlist를 보여줄 queryset 특정
        # session에 저장되어 있는 email과 room의 maker가 같은 것만 queryset에 넣음
        QuerySet = Enrol.objects.filter(
            email=self.request.session.get('user_email')).order_by('-make_date')
        return QuerySet


def classDetail_s(request, pk):
    enrol = Enrol.objects.get(pk=pk)
    room = Room.objects.get(room_id=enrol.room_id)
    request.session['room_id'] = room.room_id
    res_data = {}
    res_data['enrol_pk'] = enrol.pk
    res_data['room_name'] = room.room_name

    page = request.GET.get("page",1)
    notice_list = models.Notice.objects.filter(room_id = room.room_id).order_by('-make_date')
    paginator = Paginator(notice_list,100,orphans=5)
    try:
        notice = paginator.page(int(page))
    except EmptyPage:
        pass
    res_data["page"] = notice
    
    if request.method == 'GET':
        return render(request, 'classDetail-s.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail-s.html', res_data)


def classDetail_Detail_s(request,pk,pkk):
    enrol = Enrol.objects.get(pk=pk)
    room = Room.objects.get(room_id = enrol.room_id)
    request.session['room_id'] = room.room_id
    res_data = {}
    res_data['enrol_pk'] = enrol.pk
    res_data['room_name'] = room.room_name

    notice = Notice.objects.get(pk =pkk, room_id = room.room_id)
    res_data['title'] = notice.title
    res_data['description'] = notice.description
    res_data['writername'] = notice.writername
    res_data['date'] = notice.make_date
    
    if request.method == 'GET':
        return render(request,'classDetail-Detail-s.html',res_data)
    elif request.method == 'POST':
        return  render(request,'classDetail-Detail-s.html',res_data)

def classDetail2_s(request):
    room_session = request.session.get('room_id')
    user_session = request.session.get('user')
    user = User.objects.get(pk=user_session)
    room = Room.objects.get(room_id=room_session)
    enrol = Enrol.objects.get(room_id=room_session)
    res_data = {}
    res_data['enrol_pk'] = enrol.pk
    res_data['room_name'] = room.room_name

    page = request.GET.get("page",1)
    solvedQuiz_list = models.SolvedQuiz.objects.filter(room_id = room.room_id,user = user.email).order_by('-make_date')
    paginator = Paginator(solvedQuiz_list,100,orphans=5)
    try:
        solvedQuiz = paginator.page(int(page))
    except EmptyPage:
        pass
    res_data["page"] = solvedQuiz



    if request.method == 'GET':
        return render(request, 'classDetail2-s.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail2-s.html', res_data)

def classDetail2_Detail2_s(request,pk):
    room_session = request.session.get('room_id')
    room = Room.objects.get(room_id=room_session)
    enrol = Enrol.objects.get(room_id=room_session)
    res_data = {}
    res_data['enrol_pk'] = enrol.pk
    res_data['room_name'] = room.room_name

    quiz = SolvedQuiz.objects.get(pk =pk, room_id = room.room_id)
    res_data['makername'] = quiz.makername
    res_data['question'] = quiz.question
    res_data['item1'] = quiz.item1
    res_data['item2'] = quiz.item2
    res_data['item3'] = quiz.item3
    res_data['item4'] = quiz.item4
    res_data['answer'] = quiz.answer
    res_data['date'] = quiz.make_date

    if request.method == 'GET':
        return render(request, 'classDetail2-Detail2-s.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail2-Detail2-s.html', res_data)

def classDetail3_s(request):
    room_session = request.session.get('room_id')
    user_session = request.session.get('user')
    user = User.objects.get(pk=user_session)
    room = Room.objects.get(room_id=room_session)
    enrol = Enrol.objects.get(room_id=room_session)
    res_data = {}
    res_data['enrol_pk'] = enrol.pk
    res_data['room_name'] = room.room_name

    page = request.GET.get("page", 1)
    print(user.email, enrol.room_id, "!!!!!!!!!!!!!!!!!!!!")
    analytics_list = models.Analytics.objects.filter(
        email=user.email, room_id=enrol.room_id).order_by('-make_date')
    paginator = Paginator(analytics_list, 10, orphans=5)
    try:
        analytics = paginator.page(int(page))
    except EmptyPage:
        pass
    res_data["page"] = analytics

    if request.method == 'GET':
        return render(request, 'classDetail3-s.html', res_data)
    elif request.method == 'POST':
        return render(request, 'classDetail3-s.html', res_data)


def analyticsDetail(request, pk):
    analytics = Analytics.objects.get(pk=pk)
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role
        res_data['analyticUser'] = analytics.username

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        # chartdata 선언
        dataSource = OrderedDict()
        dataSource["data"] = []  # chartdata는 json형식이다.
        dataSource["data"].append({"label": '앱 차단', "value": analytics.app})
        dataSource["data"].append({"label": '자리이탈', "value": analytics.person})
        dataSource["data"].append({"label": '퀴즈', "value": analytics.quiz})

        chartConfig = OrderedDict()
        chartConfig["caption"] = "집중도 통계"
        chartConfig["yAxisName"] = "점수"
        chartConfig["numberSuffix"] = "점"  # y축 숫자단위
        chartConfig["theme"] = "fusion"  # 테마

        dataSource["chart"] = chartConfig  # 그래프 특징 설정

        column2D = FusionCharts(
            "column2d", "myFirstChart", "500", "400", "chart-1", "json", dataSource)
        res_data['output'] = column2D.render()

        # res_data['count'] = analytics.count
        # res_data['rate'] = analytics.rate
        res_data['level'] = analytics.level
        if request.method == 'GET':
            return render(request, 'analyticsDetail.html', res_data)
        elif request.method == 'POST':
            return render(request, 'analyticsDetail.html', res_data)
    else:
        return redirect('/login')


def classOut_t(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'roomout-success.html', res_data)
        elif request.method == 'POST':
            return redirect('/home')


def classOut_s(request):
    res_data = {}
    fs = FileSystemStorage()
    room_session = request.session.get('room_id')
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role

        room = Room.objects.get(room_id=room_session)
        users = User.objects.get(email=room.maker)
        res_data['maker'] = users.username
        res_data['maker']

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'roomout.html', res_data)
        elif request.method == 'POST':
            if user.check == False:
                return redirect('/home/class/makequiz')
            else:
                res_data['check'] = "차단 해제가 완료되지 않았습니다."
                return render(request, 'roomout.html', res_data)
    else:
        return redirect('/login')


def analytics(request,quiz):
    res_data={}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    room_session = request.session.get('room_id')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role
        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        analytics = Analytics.objects.filter(email = user.email,room_id = room_session).last() # 해당 class의 user의 최근 통계자료




        # level 판별
        appPoint = analytics.app
        personPoint = analytics.person
        quizPoint = 100- (quiz *10) 

        if (appPoint + personPoint + quizPoint >= 299):
            level = 10
            list = 1
        elif (appPoint + personPoint + quizPoint >= 270 and appPoint + personPoint + quizPoint < 299):
            level = 9
            list = 1
        elif (appPoint + personPoint + quizPoint >= 240 and appPoint + personPoint + quizPoint < 270):
            level = 8
            list = 2
        elif (appPoint + personPoint + quizPoint >= 210 and appPoint + personPoint + quizPoint < 240):
            level = 7
            list = 2
        elif (appPoint + personPoint + quizPoint >= 180 and appPoint + personPoint + quizPoint < 210):
            level = 6
            list = 3
        elif (appPoint + personPoint + quizPoint >= 150 and appPoint + personPoint + quizPoint < 180):
            level = 5
            list = 3
        elif (appPoint + personPoint + quizPoint >= 120 and appPoint + personPoint + quizPoint < 150):
            level = 4
            list = 3
        elif (appPoint + personPoint + quizPoint >= 90 and appPoint + personPoint + quizPoint < 120):
            level = 3
            list = 3
        elif (appPoint + personPoint + quizPoint >= 60 and appPoint + personPoint + quizPoint < 90):
            level = 2
            list = 3
        elif (appPoint + personPoint + quizPoint < 60 and appPoint + personPoint + quizPoint >= 30):
            level = 1
            list = 3
        elif (appPoint + personPoint + quizPoint == 0 ):
            level = 0
            list = 3    
        else:
            level = 0
            list = 3

        analytics.level = level
        analytics.list = list
        analytics.quiz = quizPoint
        analytics.save()


        # # 2. rate 판별  (모두 다 나왔을때, 나오지 않았을때 고려 해야함)
        # room_name = request.session.get('room_name')
        # analytic = Analytics.objects.filter(room_name = room_name).order_by('-level')  #세션에 있는 room_name과 같은 통계 자료를 가져옴
        # num = 0
        # for x in analytic:
        #     if x.email == user.email:
        #         x.rate = num + 1
        #         x.save()
        #     else:
        #         num = num+1

        dataSource = OrderedDict()
        dataSource["data"] = [] #chartdata는 json형식이다.
        dataSource["data"].append({"label": '앱 차단', "value": appPoint})
        dataSource["data"].append({"label": '자리이탈', "value": personPoint})
        dataSource["data"].append({"label": '퀴즈점수', "value": quizPoint})

        chartConfig = OrderedDict()
        chartConfig["caption"] = "집중도 통계"  
        chartConfig["numberSuffix"] = "점" #y축 숫자단위
        chartConfig["theme"] = "fusion" #테마

        dataSource["chart"] = chartConfig # 그래프 특징 설정

        column2D = FusionCharts("column2d", "myFirstChart", "500", "400", "chart-1", "json", dataSource)
        res_data['output'] = column2D.render() 

        # analytics = Analytics.objects.filter(email = user.email).last()
        # res_data['count'] = Analytics.objects.filter(room_name = room_name).count()
        # analytics.count = res_data['count']
        # analytics.save()
        # res_data['rate'] = analytics.rate
        res_data['level'] = analytics.level
        res_data['analyticUser'] = analytics.username
        del(request.session['room_id'])  # Class가 이제 끝났으니 room session 삭제!!!!
        if request.method == 'GET':
            return render(request,'analyticsDetail.html',res_data)
        elif request.method == 'POST':  # 집중도에 사용할 데이터 받는 POST
            return  render(request,'analyticsDetail.html',res_data)
    else:
        return redirect('/login')


@csrf_exempt
def student_quiz(request):

    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        # room을 만든 사람 username 가져오기
        room_session = request.session.get('room_id')
        room = Room.objects.get(room_id=room_session)
        room_maker = User.objects.get(email=room.maker)
        res_data['room_maker'] = room_maker.username

        if request.method == 'GET':

            # 룸ID 받아오기
            # room_session = request.session.get('room_id')
            # print(room_session)
            past_class = Enrol.objects.filter(
                email=user.email).order_by('-make_date')
            print(past_class)
            quiz = Quiz.objects.filter(room_id=past_class[0].room_id)
            print(quiz)
            # 학생 복습퀴즈시 자신이 낸 문제는 제외

            except_quiz = []
            for i in quiz:
                if(i.maker != user.email and i.maker != room_maker):
                    except_quiz.append(i)

            print(except_quiz)

            random_num = random.randint(1, len(except_quiz))

            print(random_num)

            maker = except_quiz[random_num-1].maker
            quiz_ower = User.objects.get(email=maker)
            room_id = except_quiz[random_num-1].room_id
            room = Room.objects.get(room_id=room_id)
            res_data['question'] = except_quiz[random_num-1].question
            res_data['item1'] = except_quiz[random_num-1].item1
            res_data['item2'] = except_quiz[random_num-1].item2
            res_data['item3'] = except_quiz[random_num-1].item3
            res_data['item4'] = except_quiz[random_num-1].item4
            res_data['maker'] = quiz_ower.username
            res_data['id'] = except_quiz[random_num-1].id
            res_data['room_name'] = room.room_name
            print(room.room_name)
            return render(request, 'quiz.html', res_data)
        elif request.method == 'POST':
            info = {}
            id = request.POST.get('id')
            answer = request.POST.get('check')
            print(id)
            try:
                quiz = Quiz.objects.get(id=id)
            except Quiz.DoesNotExist:
                return messages.warning(request, '존재 하는 Class가 없습니다.')


            try:
                solvedQuiz = SolvedQuiz.objects.get(id=quiz.pk,user= user.email)  # 그 문제를 이 user가 풀었는지 보고
            except SolvedQuiz.DoesNotExist:                                  # 안 풀었을 때만 저장
                solvedQuiz = SolvedQuiz(id = id, user = user.email,makername= quiz.makername,room_id = room.room_id, question = quiz.question,
                 item1 = quiz.item1, item2 = quiz.item2, item3 = quiz.item3, item4 = quiz.item4, answer = quiz.answer)
                solvedQuiz.save()


            quiz_answer = quiz.answer
            print(quiz_answer)

            if answer == str(quiz_answer):
                info['result'] = "yes"
            elif answer != str(quiz_answer):
                info['answer'] = quiz_answer
                info['result'] = "no"


            solvedQuiz = SolvedQuiz(id = id, user = user.email,room_id = room.room_id)
            solvedQuiz.save()

            return JsonResponse(info)

    else:
        return redirect('/login')


@csrf_exempt
def teacher_quiz(request):

    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')

    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        # room을 만든 사람 username 가져오기
        room_session = request.session.get('room_id')
        room = Room.objects.get(room_id=room_session)
        room_maker = User.objects.get(email=room.maker)
        res_data['room_maker'] = room_maker.username
        # 룸ID 받아오기
            # room_session = request.session.get('room_id')
            # print(room_session)

        quiz_all = Quiz.objects.filter(room_id=room.room_id)
        print(quiz_all)
        teacher_quiz = []
        for i in quiz_all:
            if(i.maker==room_maker.email):
                teacher_quiz.append(i)

        res_data['question'] = teacher_quiz[0].question
        res_data['item1'] = teacher_quiz[0].item1
        res_data['item2'] = teacher_quiz[0].item2
        res_data['item3'] = teacher_quiz[0].item3
        res_data['item4'] = teacher_quiz[0].item4
        res_data['maker'] = "선생님"
        res_data['id'] = teacher_quiz[0].id
        res_data['room_name'] = room.room_name
        print(room.room_name)

        if request.method == 'GET':
            return render(request, 'teacherquiz.html', res_data)
        elif request.method == 'POST':
            answer = request.POST.get('check')
            id = request.POST.get('id')
            info = {}
            print(id)
            try:
                quiz = Quiz.objects.get(id=id)
            except Quiz.DoesNotExist:
                return messages.warning(request, '존재 하는 Class가 없습니다.')


            try:
                solvedQuiz = SolvedQuiz.objects.get(id=quiz.pk,user= user.email)  # 그 문제를 이 user가 풀었는지 보고
            except SolvedQuiz.DoesNotExist:                                  # 안 풀었을 때만 저장
                solvedQuiz = SolvedQuiz(id = id, user = user.email,makername= quiz.makername,room_id = room.room_id, question = quiz.question,
                 item1 = quiz.item1, item2 = quiz.item2, item3 = quiz.item3, item4 = quiz.item4, answer = quiz.answer)
                solvedQuiz.save()

            quiz_answer = quiz.answer
            print(quiz_answer)

            if answer == str(quiz_answer):
                info['result'] = "yes"
            elif answer != str(quiz_answer):
                info['answer'] = quiz_answer
                info['result'] = "no"
            return JsonResponse(info)

    else:
        return redirect('/login')


@csrf_exempt
def make_quiz(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')

    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        res_data['role'] = user.role

        # room을 만든 사람 username 가져오기
        room_session = request.session.get('room_id')
        room = Room.objects.get(room_id=room_session)
        room_maker = User.objects.get(email=room.maker)
        res_data['room_maker'] = room_maker.username

        if request.method == 'GET':

            return render(request, 'makequiz.html', res_data)

        elif request.method == 'POST':
            question = request.POST.get('question')
            item1 = request.POST.get('item1')
            item2 = request.POST.get('item2')
            item3 = request.POST.get('item3')
            item4 = request.POST.get('item4')
            answer = request.POST.get('answer')
            print(question)

            # 나의 기존 퀴즈 삭제
            quiz_all = Quiz.objects.filter(room_id=room.room_id)
            id="NULL"
            for i in quiz_all:
                if(i.maker == user.email):
                    id=i.id
            if(id!="NULL"):
                myquiz = Quiz.objects.get(id=id)
                myquiz.delete()
            
            quiz = Quiz(room_id=room.room_id, maker=user.email, makername = user.username, question=question,
                        item1=item1, item2=item2, item3=item3, item4=item4, answer=answer)
            quiz.save()

            if(user.email == room_maker.email):
                res_data['result'] = "teacher"
            else:
                res_data['result'] = "student"

            return JsonResponse(res_data)
    else:
        return redirect('/login')


@method_decorator(csrf_exempt, name='dispatch')
def app_enterroom(request):
    # 앱에서 오는 로그인 요청
    if request.method == "POST":
        roomid = request.POST.get('roomname', None)
        password = request.POST.get('password', None)
        # 받은 이메일이랑 비밀번호 =데이터와 일치하면
        # 리턴값으로 숫자 200 = 로그인 성공
        # 일치 안하면 숫자 100 = 로그인 실패
        if Room.objects.filter(room_id=roomid).exists():
            room = Room.objects.get(room_id=roomid)
            # db에서 꺼내는 명령.Post로 받아온 username으로 , db의 username을 꺼내온다.
            print(password)
            print(room.room_id)
            if password == room.room_password:
                print(room.room_password)
                # 세션도 딕셔너리 변수 사용과 똑같이 사용하면 된다.
                # 세션 user라는 key에 방금 로그인한 id를 저장한것.
                # room_type(exam mode, study mode)구분해서 들고오기
                return HttpResponse("success")
            else:
                # 방 비번 불일치
                return HttpResponse("fail")

        else:
            # 방 없음
            return HttpResponse("NoRoom")


@method_decorator(csrf_exempt, name='dispatch')
def app_attendance(request):
    if request.method == "POST":
        info = {}
        # change_here
        room_id = request.POST.get('room', None)
        member_name = request.POST.get('name', None)
        member_number = request.POST.get('number', None)

        print(room_id)
        print(member_name)
        print(member_number)
        # room DB-member_list로 회원번호 확인 및 index 추출
        room = Room.objects.get(room_id=room_id)
        # member_list = room.member_list  # 회원번호만 적힌 리스트
        # member_list = member_list[1:-1].split(', ')
        # print(member_list)

        # CHECK NUMBER
        # Correct NUMBER
        # if (member_number in member_list):
        #     member_index = member_list.index(member_number) + 1
        #     print('member_index:'+str(member_index))

        #     # 해당방의 DB속 명단Excel파일 조회
        #     room = Room.objects.get(room_name=room_name)
        #     member_file = room.file  # 명단

        #     member = load_workbook("media/" + str(member_file))
        #     sheet = member['Sheet1']
        #     member_file_name = sheet['B'+str(member_index)].value

        #     # CHECK NAME
        #     # Wrong NAME
        #     if member_file_name != member_name:
        #         print('app_enterEXAM_info_no_match_name_num')
        #         return HttpResponse(simplejson.dumps({"roomname": "no",  "password": "no"}))
        #     # Correct NAME
        #     else:
        #         print('app_enterEXAM_info_success')
        #         return HttpResponse(simplejson.dumps({"roomname": "yes", "password": member_index}))

        # # Wrong NUMBER
        # else:
        #     print('app_enterEXAM_info_no_num')
        #     return HttpResponse(simplejson.dumps({"roomname": "fail", "password": "no"}))

        return HttpResponse("success")


@method_decorator(csrf_exempt, name='dispatch')
def app_checkimg(request):
    if request.method == "POST":
        # APP : 캡쳐이미지 받기
        capture_image = request.FILES['image']
        print(capture_image)
        # image.name = <room_name>_<member_number>_capture.png
        fs = FileSystemStorage()
        filename = fs.save("capture/"+capture_image.name, capture_image)
        uploaded_file_url = fs.url(filename)
        print(filename)
        print(uploaded_file_url)

        # image.name 에서 분리
        l = capture_image.name.split('_')
        room_name = l[0]
        member_index = l[1]

        print("Get All DATA ")

        # # Room DB - excel 파일
        # room = Room.objects.get(room_name=room_name)
        # member_file = room.file  # 명단
        # member = load_workbook("media/" + str(member_file))
        # sheet = member['Sheet1']

        # # exel 명단 속 이미지
        # image_loader = SheetImageLoader(sheet)
        # image = image_loader.get('C'+str(member_index))
        # member_file_image_path = (room_name+"_"+str(member_index)+".jpg")
        # image.save("media/capture/"+member_file_image_path)
        # fs = FileSystemStorage()

        # # Face Recognition
        # a = (fs.location + str("/capture/") + member_file_image_path)
        # b = (fs.location + str("/capture/") + capture_image.name)
        # # luxand API
        # luxand_client = luxand("12a42a8efedf4e24b84730ce440e5429")
        # member_file_image = luxand_client.add_person(
        #     str(member_index), photos=[a])
        # result = luxand_client.verify(member_file_image, photo=b)
        # print(result)
        # os.remove(os.path.join(settings.MEDIA_ROOT +
        #                        "/capture/", capture_image.name))
        # # Recognition RESULT
        # if result['status'] == 'success':
        #     print("Recognition_SUCCESS")
        #     return HttpResponse(simplejson.dumps({"image": "ok"}))  # 이미지 전송완료
        # else:
        #     print("Recognition_FAIL")
        #     return HttpResponse(simplejson.dumps({"image": "no"}))  # 이미지 전송실패

        return HttpResponse("success")


@method_decorator(csrf_exempt, name='dispatch')
def app_sendcount(request):
    if request.method == "POST":
        email = request.POST.get('email', None)
        app = request.POST.get('count', None)  # 앱 접근횟수
        nonperson = request.POST.get('nonperson', None)  # 자리이탈횟수
        roomname = request.POST.get('roomname', None)
        print(app)
        app_point = int(100) - (int(app)*10)
        nonperson_point = int(100)-(int(nonperson)*5)
        output = ''
        print(nonperson)
        print(roomname)

        myuser = User.objects.get(email=email)
        myuser.check = False
        myuser.save()

        analytics = Analytics(room_id=roomname, email=email,username=myuser.username, person=nonperson_point, app=app_point)
        analytics.save()

        # analytics = Analytics.objects.filter(room_name=roomname)
        # if analytics:  # 해당 룸의 row가 있다.
        #     for x in analytics:
        #         if(x.email == myuser.email):   # 해당 룸의 내 email을 가진 row가 있다.
        #             output = 'YES'
        #         else:
        #             output = 'NO'
        # elif not(analytics):  # 해당 룸의 row가 없다 -> 내가 제일 처음 -> 바로 생성
        #     analytics = Analytics(
        #         room_id=roomname, email=email, person=nonperson_point, app=app_point)
        #     analytics.save()

        # if output == 'YES':  # 해당 룸의 row 중에 내 아이디의 row가 있다.
        #     analytics = Analytics.objects.filter(email=email).last()
        #     analytics.person = nonperson_point
        #     analytics.app = count_point
        #     analytics.save()
        # elif output == 'NO':  # 해당 룸의 row 중에 내 아이디의 row가 없다.
        #     analytics = Analytics(
        #         room_name=roomname, email=email, person=nonperson_point, app=count_point)
        #     analytics.save()

        return HttpResponse("success")
